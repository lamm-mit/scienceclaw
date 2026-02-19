#!/usr/bin/env python3
"""
XLSX/CSV extraction tool - extract and preview spreadsheet data from
Excel and CSV files for scientific supplementary data analysis.
"""

import argparse
import json
import os
import sys


def extract_xlsx_openpyxl(file_path: str, sheet_name: str, head: int) -> dict:
    """Extract from .xlsx using openpyxl."""
    import openpyxl

    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    sheet_names = wb.sheetnames

    target_sheets = [sheet_name] if sheet_name else sheet_names

    data = {}
    total_rows = 0
    total_cols = 0

    for sname in target_sheets:
        if sname not in sheet_names:
            continue
        ws = wb[sname]
        rows = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i >= head:
                break
            rows.append([str(cell) if cell is not None else "" for cell in row])
        data[sname] = rows
        # Shape from the full sheet
        total_rows = max(total_rows, ws.max_row or 0)
        total_cols = max(total_cols, ws.max_column or 0)

    wb.close()

    return {
        "file": file_path,
        "sheets": sheet_names,
        "data": data,
        "shape": {"rows": total_rows, "cols": total_cols},
        "extractor": "openpyxl",
    }


def extract_csv_pandas(file_path: str, head: int) -> dict:
    """Extract from .csv using pandas."""
    import pandas as pd

    df = pd.read_csv(file_path)
    preview = df.head(head)

    # Convert to list of lists with header
    rows = [list(preview.columns)] + preview.values.tolist()
    rows = [[str(cell) for cell in row] for row in rows]

    sheet_name = os.path.basename(file_path)
    return {
        "file": file_path,
        "sheets": [sheet_name],
        "data": {sheet_name: rows},
        "shape": {"rows": len(df), "cols": len(df.columns)},
        "extractor": "pandas",
    }


def extract_xlsx_pandas(file_path: str, sheet_name: str, head: int) -> dict:
    """Extract from .xlsx using pandas as fallback."""
    import pandas as pd

    xl = pd.ExcelFile(file_path)
    sheet_names = xl.sheet_names

    target_sheets = [sheet_name] if sheet_name else sheet_names

    data = {}
    total_rows = 0
    total_cols = 0

    for sname in target_sheets:
        if sname not in sheet_names:
            continue
        df = pd.read_excel(file_path, sheet_name=sname)
        preview = df.head(head)
        rows = [list(preview.columns)] + preview.values.tolist()
        rows = [[str(cell) for cell in row] for row in rows]
        data[sname] = rows
        total_rows = max(total_rows, len(df))
        total_cols = max(total_cols, len(df.columns))

    return {
        "file": file_path,
        "sheets": sheet_names,
        "data": data,
        "shape": {"rows": total_rows, "cols": total_cols},
        "extractor": "pandas",
    }


def main():
    parser = argparse.ArgumentParser(
        description="Extract and preview data from Excel and CSV spreadsheet files"
    )
    parser.add_argument("--file", required=True, help="Path to .xlsx, .xls, or .csv file")
    parser.add_argument("--sheet", default=None, help="Sheet name to extract (default: all sheets)")
    parser.add_argument(
        "--head",
        type=int,
        default=20,
        help="Number of rows to preview per sheet (default: 20)",
    )
    args = parser.parse_args()

    if not os.path.isfile(args.file):
        result = {"error": f"File not found: {args.file}", "file": args.file}
        print(json.dumps(result, indent=2))
        sys.exit(1)

    ext = os.path.splitext(args.file)[1].lower()

    try:
        if ext == ".csv":
            try:
                import pandas as pd  # noqa: F401
                result = extract_csv_pandas(args.file, args.head)
            except ImportError:
                result = {
                    "error": "pandas not installed for CSV processing. Run: pip install pandas",
                    "file": args.file,
                }
                print(json.dumps(result, indent=2))
                sys.exit(1)

        elif ext in (".xlsx", ".xls"):
            try:
                import openpyxl  # noqa: F401
                result = extract_xlsx_openpyxl(args.file, args.sheet, args.head)
            except ImportError:
                try:
                    import pandas as pd  # noqa: F401
                    result = extract_xlsx_pandas(args.file, args.sheet, args.head)
                except ImportError:
                    result = {
                        "error": (
                            "Neither openpyxl nor pandas installed. "
                            "Run: pip install openpyxl pandas"
                        ),
                        "file": args.file,
                    }
                    print(json.dumps(result, indent=2))
                    sys.exit(1)
        else:
            result = {
                "error": f"Unsupported file type '{ext}'. Supported: .xlsx, .xls, .csv",
                "file": args.file,
            }
            print(json.dumps(result, indent=2))
            sys.exit(1)

    except Exception as e:
        result = {"error": str(e), "file": args.file}
        print(json.dumps(result, indent=2))
        sys.exit(1)

    print(json.dumps(result, indent=2))


if __name__ == "__main__":
    main()
